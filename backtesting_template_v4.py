"""
Script para generar plantilla Excel de backtesting automática
Uso: python backtesting_template_v4.py
Genera: backtesting_v4.xlsx
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

def create_backtesting_template():
    """Crea plantilla Excel profesional para backtesting v4.0"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Backtesting v4.0"
    
    # ─────────────────────────────────────────────────────────────────────
    # ENCABEZADO
    # ─────────────────────────────────────────────────────────────────────
    ws['A1'] = "BACKTESTING DETALLADO - BOT v4.0"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:O1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A2'] = f"Fecha: {datetime.date.today()}"
    ws['A2'].font = Font(size=10)
    ws.merge_cells('A2:C2')
    
    # ─────────────────────────────────────────────────────────────────────
    # INSTRUCCIONES
    # ─────────────────────────────────────────────────────────────────────
    ws['A4'] = "INSTRUCCIONES:"
    ws['A4'].font = Font(size=11, bold=True, color="C00000")
    
    instructions = [
        "1. Completa cada fila con UN TRADE que tu bot habría abierto",
        "2. Usa TradingView Replay para simular en tiempo real",
        "3. Registra: Tier, Entry, SL, TP1, TP2, MFI, ZLSMA, ADX, Precio de salida",
        "4. Resultado = (Salida - Entry) / (Entry - SL) en unidades R",
        "5. Las métricas se calculan automáticamente abajo",
        "6. Target: Win Rate ≥65%, Expectancy ≥+0.5R"
    ]
    
    for i, instruction in enumerate(instructions, start=5):
        ws[f'A{i}'] = instruction
        ws[f'A{i}'].font = Font(size=9, italic=True)
    
    # ─────────────────────────────────────────────────────────────────────
    # ENCABEZADOS DE COLUMNAS
    # ─────────────────────────────────────────────────────────────────────
    headers = [
        "Fecha", "Hora", "Par", "Tier", "Direction", "Entry", "SL", "TP1", "TP2",
        "MFI", "ZLSMA %", "ADX", "Salida Real", "Resultado (R)", "Win", "Notas"
    ]
    
    header_row = 12
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Ajusta altura del encabezado
    ws.row_dimensions[header_row].height = 25
    
    # ─────────────────────────────────────────────────────────────────────
    # FILA DE EJEMPLO
    # ─────────────────────────────────────────────────────────────────────
    example_row = 13
    example_data = [
        "2024-01-15",
        "15:45",
        "BTC-USDT",
        "A",
        "LONG",
        42500,
        42350,
        42650,
        42800,
        35,
        "+0.8",
        28,
        42750,
        "=IF(D13=\"LONG\",(M13-F13)/(F13-G13),(H13-M13)/(H13-G13))",
        "=IF(N13>0,1,0)",
        "Clean entry on ZLSMA bounce"
    ]
    
    for col, value in enumerate(example_data, start=1):
        cell = ws.cell(row=example_row, column=col)
        cell.value = value
        if col in [1, 2]:  # Fecha y hora
            cell.number_format = 'dd-mm-yyyy' if col == 1 else 'hh:mm'
        elif col in [6, 7, 8, 9, 13]:  # Precios
            cell.number_format = '0.00'
        elif col in [10, 11, 12]:  # Indicadores
            cell.number_format = '0.0'
        elif col == 14:  # Resultado
            cell.number_format = '0.00'
        cell.alignment = Alignment(horizontal='center')
        
        # Resalta ejemplo en gris claro
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # ─────────────────────────────────────────────────────────────────────
    # FILAS DE ENTRADA (14-113 = 100 filas)
    # ─────────────────────────────────────────────────────────────────────
    for row in range(14, 114):
        # Crear fórmulas para cálculo automático
        ws.cell(row=row, column=14).value = f"=IF(AND(F{row}<>\"\",G{row}<>\"\",M{row}<>\"\"),IF(D{row}=\"LONG\",(M{row}-F{row})/(F{row}-G{row}),(H{row}-M{row})/(H{row}-G{row})),\"\")"
        ws.cell(row=row, column=14).number_format = '0.00'
        ws.cell(row=row, column=15).value = f"=IF(N{row}>0,1,IF(N{row}<0,0,\"\"))"
        ws.cell(row=row, column=15).number_format = '0'
        
        # Formato de precio
        for col in [6, 7, 8, 9, 13]:
            ws.cell(row=row, column=col).number_format = '0.00'
        
        # Formato de porcentaje
        for col in [10, 11, 12]:
            ws.cell(row=row, column=col).number_format = '0.0'
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for col in range(1, 17):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    
    # ─────────────────────────────────────────────────────────────────────
    # SECCIÓN DE MÉTRICAS
    # ─────────────────────────────────────────────────────────────────────
    metrics_row = 116
    
    ws[f'A{metrics_row}'] = "MÉTRICAS CALCULADAS"
    ws[f'A{metrics_row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{metrics_row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells(f'A{metrics_row}:C{metrics_row}')
    
    metrics = [
        ("Total Trades", f"=COUNTA(F14:F113)", metrics_row + 2),
        ("Win Rate %", f"=IF(B{metrics_row+2}=0,0,(COUNTIF(O14:O113,1)/B{metrics_row+2})*100)", metrics_row + 2),
        ("Trades Ganados", f"=COUNTIF(O14:O113,1)", metrics_row + 2),
        ("Trades Perdidos", f"=COUNTIF(O14:O113,0)", metrics_row + 3),
        
        ("Tier S Trades", f'=COUNTIF(D14:D113,"S")', metrics_row + 4),
        ("Tier S Win %", f'=IF(E{metrics_row+4}=0,0,(SUMIFS(O14:O113,D14:D113,"S")/E{metrics_row+4})*100)', metrics_row + 4),
        
        ("Tier A Trades", f'=COUNTIF(D14:D113,"A")', metrics_row + 5),
        ("Tier A Win %", f'=IF(F{metrics_row+5}=0,0,(SUMIFS(O14:O113,D14:D113,"A")/F{metrics_row+5})*100)', metrics_row + 5),
        
        ("Promedio R Ganadores", f"=IF(COUNTIF(O14:O113,1)=0,0,AVERAGEIF(O14:O113,\">0\",N14:N113))", metrics_row + 6),
        ("Promedio R Perdedores", f"=IF(COUNTIF(O14:O113,0)=0,0,ABS(AVERAGEIF(O14:O113,\"<0\",N14:N113)))", metrics_row + 6),
        
        ("Profit Factor", f"=IF(SUMIF(N14:N113,\"<0\",N14:N113)=0,9999,SUMIF(N14:N113,\">0\",N14:N113)/ABS(SUMIF(N14:N113,\"<0\",N14:N113)))", metrics_row + 7),
        ("Total R Ganados", f"=SUMIF(N14:N113,\">0\",N14:N113)", metrics_row + 7),
        ("Total R Perdidos", f"=ABS(SUMIF(N14:N113,\"<0\",N14:N113))", metrics_row + 7),
        ("Net R", f"=H{metrics_row+7}-I{metrics_row+7}", metrics_row + 7),
        
        ("Expectancy", f"=(B{metrics_row+2}/B{metrics_row}*G{metrics_row+6})-(D{metrics_row+3}/B{metrics_row}*H{metrics_row+6})", metrics_row + 8),
        ("Max Drawdown", "= (ver nota abajo)", metrics_row + 8),
    ]
    
    col_metric = 1
    col_value = 3
    col_note = 5
    
    for i, (metric_name, formula, row) in enumerate(metrics):
        # Nombre de métrica
        cell_name = ws.cell(row=row, column=col_metric)
        cell_name.value = metric_name
        cell_name.font = Font(bold=True, size=10)
        cell_name.fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
        
        # Fórmula
        cell_value = ws.cell(row=row, column=col_value)
        cell_value.value = formula
        cell_value.font = Font(size=10)
        cell_value.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        cell_value.number_format = '0.00'
        
        if '%' in metric_name or 'Rate' in metric_name:
            cell_value.number_format = '0.0"%"'
    
    # ─────────────────────────────────────────────────────────────────────
    # NOTAS FINALES
    # ─────────────────────────────────────────────────────────────────────
    note_row = metrics_row + 11
    
    ws[f'A{note_row}'] = "INTERPRETACIÓN DE MÉTRICAS"
    ws[f'A{note_row}'].font = Font(size=11, bold=True, color="C00000")
    
    notes = [
        "✅ Win Rate ≥65%: Bueno | ≥72%: Excelente | <60%: Revisar estrategia",
        "✅ Tier S Win % ≥75%: Filtro está funcionando bien",
        "✅ Tier A Win % ≥65%: Aceptable",
        "✅ Profit Factor ≥2.0: Excelente | 1.5-2.0: Bueno | <1.5: Revisar",
        "✅ Expectancy ≥+0.5R: Rentable | ≥+1.0R: Muy rentable",
        "✅ Max Drawdown <15%: Aceptable | <10%: Bueno (cuenta pérdidas consecutivas)",
        "",
        "FÓRMULA DE EXPECTANCY:",
        "E = (Win% × Avg_Win_R) - (Loss% × Avg_Loss_R)",
        "Si E > +0.5R, la estrategia es rentable a largo plazo",
        "",
        "CONVERSIÓN A DINERO REAL:",
        "Si E = +1.0R y riesgas $10 por trade:",
        "Ganancia mensual esperada = $10 × E × trades/mes",
        "Ejemplo: $10 × 1.0R × 40 trades = $400/mes (con comisiones: ~$350)",
    ]
    
    for i, note in enumerate(notes, start=1):
        cell = ws.cell(row=note_row + i, column=1)
        cell.value = note
        cell.font = Font(size=9, italic=True if not note.startswith('✅') else False)
        ws.merge_cells(f'A{note_row+i}:F{note_row+i}')
    
    # ─────────────────────────────────────────────────────────────────────
    # AJUSTES DE COLUMNAS
    # ─────────────────────────────────────────────────────────────────────
    column_widths = {
        'A': 12, 'B': 10, 'C': 12, 'D': 8, 'E': 11,
        'F': 11, 'G': 11, 'H': 11, 'I': 11, 'J': 8,
        'K': 10, 'L': 8, 'M': 12, 'N': 13, 'O': 10, 'P': 20
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # ─────────────────────────────────────────────────────────────────────
    # SHEET 2: Resumen Tier S/A/B
    # ─────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen por Tier")
    
    ws2['A1'] = "ANÁLISIS DETALLADO POR TIER"
    ws2['A1'].font = Font(size=12, bold=True, color="FFFFFF")
    ws2['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws2.merge_cells('A1:F1')
    
    ws2['A3'] = "Tier"
    ws2['B3'] = "Total Trades"
    ws2['C3'] = "Ganados"
    ws2['D3'] = "Perdidos"
    ws2['E3'] = "Win %"
    ws2['F3'] = "Avg R"
    
    for col in range(1, 7):
        cell = ws2.cell(row=3, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Datos para cada tier
    tiers = ['S', 'A', 'B']
    for idx, tier in enumerate(tiers, start=4):
        ws2.cell(row=idx, column=1).value = tier
        ws2.cell(row=idx, column=2).value = f'=COUNTIF(Backtesting!D14:D113,"{tier}")'
        ws2.cell(row=idx, column=3).value = f'=SUMIFS(Backtesting!O14:O113,Backtesting!D14:D113,"{tier}")'
        ws2.cell(row=idx, column=4).value = f'=B{idx}-C{idx}'
        ws2.cell(row=idx, column=5).value = f'=IF(B{idx}=0,0,(C{idx}/B{idx})*100)'
        ws2.cell(row=idx, column=6).value = f'=AVERAGEIFS(Backtesting!N14:N113,Backtesting!D14:D113,"{tier}")'
        
        ws2.cell(row=idx, column=5).number_format = '0.0"%"'
        ws2.cell(row=idx, column=6).number_format = '0.00'
        
        for col in range(1, 7):
            ws2.cell(row=idx, column=col).alignment = Alignment(horizontal='center')
    
    # Ajusta columnas sheet 2
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws2.column_dimensions[col].width = 14
    
    # Guarda el libro
    wb.save('backtesting_v4.xlsx')
    print("✅ Plantilla creada: backtesting_v4.xlsx")
    print("\nPróximos pasos:")
    print("1. Abre backtesting_v4.xlsx")
    print("2. Reemplaza la fila 13 (ejemplo) con tus datos reales")
    print("3. Completa mínimo 50 filas para estadísticas confiables")
    print("4. Las métricas se calculan automáticamente abajo")
    print("\nTarget para dinero real:")
    print("  • Win Rate ≥65%")
    print("  • Expectancy ≥+0.5R")
    print("  • Tier S Win % ≥75%")

if __name__ == "__main__":
    create_backtesting_template()
